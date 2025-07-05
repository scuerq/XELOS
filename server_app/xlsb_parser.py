import pandas as pd
from collections import defaultdict

class XLSBParser:
    """Parse XLSB files according to mappings defined in the VBA modules."""

    def __init__(self, path):
        self.path = path

    def _read_range(self, sheet_name, cell_range):
        df = pd.read_excel(self.path, sheet_name=sheet_name, engine='pyxlsb', header=None)
        # Convert Excel-style range (e.g., "F18:F28") to zero-based index slicing
        # This is simplified: we read full sheet and slice with pandas' iloc.
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        print("parse_synth columns:", df.columns)
        print(df.head())
        return df.iloc[min_row-1:max_row, min_col-1:max_col]

    def _offset_range(self, cell_range, col_offset=0, row_offset=0):
        """Return a new range string offset from the original range."""
        from openpyxl.utils import range_boundaries, get_column_letter
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        min_col += col_offset
        max_col += col_offset
        min_row += row_offset
        max_row += row_offset
        from openpyxl.utils import get_column_letter
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def parse_id(self):
        """Parse identification information from the 'Identif' sheet."""
        tables = {}
        try:
            ranges = [
                ('C7:D21'),
                ('M19:O19'),
                ('M22:O22'),
                ('G85:M85'),
                ('C86:E86'),
            ]
            frames = []
            for r in ranges:
                df = self._read_range('Identif', r)
                if df.shape[1] < 2:
                    continue
                df = df.iloc[:, :2]
                frames.append(df)
            if frames:
                id_df = pd.concat(frames, ignore_index=True)
                id_df.columns = ['Key', 'Value']
                tables['ID'] = id_df
        except Exception:
            pass
        return tables

    def compute_synthese(self, df):
        import numpy as np
        results = []
        if 'SOCIETE' not in df.columns:
            return pd.DataFrame()  # Sécurité : pas de colonne SOCIETE
        for societe, group in df.groupby('SOCIETE'):
            nb_operations = group['Id'].nunique() if 'Id' in group else np.nan
            nb_lgt = group['Equivalent logements'].sum() if 'Equivalent logements' in group else np.nan
            shab = group['Nbre de M² SHab'].sum() if 'Nbre de M² SHab' in group else np.nan
            prp = group['Total'].sum() if 'Total' in group else np.nan
            prp_shab = prp / shab if shab else np.nan
            prp_logt = prp / nb_lgt if nb_lgt else np.nan
            loyer_moyen = group['Montant'].sum() / nb_lgt / 12 if nb_lgt and 'Montant' in group else np.nan
            taux_sub = group['Subventions'].sum() / prp if prp and 'Subventions' in group else np.nan
            trlb = (group['Taux de rentabilité locative brute immédiat'] * group['Total']).sum() / prp if prp and 'Taux de rentabilité locative brute immédiat' in group and 'Total' in group else np.nan
            ebe = (group['EBE / Loyers potentiels Cumulés'] * group['Montant']).sum() / group['Montant'].sum() if 'EBE / Loyers potentiels Cumulés' in group and 'Montant' in group and group['Montant'].sum() else np.nan
            dette = (group['Service de la Dette / Loyers potentiels Cumulés'] * group['Montant']).sum() / group['Montant'].sum() if 'Service de la Dette / Loyers potentiels Cumulés' in group and 'Montant' in group and group['Montant'].sum() else np.nan
            autofi = (group["Taux d'Autofinancement"] * group['Montant']).sum() / group['Montant'].sum() if "Taux d'Autofinancement" in group and 'Montant' in group and group['Montant'].sum() else np.nan
            results.append({
                'SOCIETE': societe,
                '_NB_OPERATIONS': nb_operations,
                '_NB_LGT': nb_lgt,
                '_SHAB': shab,
                '_PRP': prp,
                '_PRP_SHAB': prp_shab,
                '_PRP_LOGT': prp_logt,
                '_LOYER_MOYEN_MENSUEL': loyer_moyen,
                '_TAUX_SUB': taux_sub,
                '_TRLB': trlb,
                '_EBE': ebe,
                '_DETTE': dette,
                '_AUTOFI': autofi,
            })
        # Total général
        total = {
            'SOCIETE': 'Total général',
            '_NB_OPERATIONS': df['Id'].nunique() if 'Id' in df else np.nan,
            '_NB_LGT': df['Equivalent logements'].sum() if 'Equivalent logements' in df else np.nan,
            '_SHAB': df['Nbre de M² SHab'].sum() if 'Nbre de M² SHab' in df else np.nan,
            '_PRP': df['Total'].sum() if 'Total' in df else np.nan,
            '_PRP_SHAB': df['Total'].sum() / df['Nbre de M² SHab'].sum() if 'Total' in df and 'Nbre de M² SHab' in df and df['Nbre de M² SHab'].sum() else np.nan,
            '_PRP_LOGT': df['Total'].sum() / df['Equivalent logements'].sum() if 'Total' in df and 'Equivalent logements' in df and df['Equivalent logements'].sum() else np.nan,
            '_LOYER_MOYEN_MENSUEL': df['Montant'].sum() / df['Equivalent logements'].sum() / 12 if 'Montant' in df and 'Equivalent logements' in df and df['Equivalent logements'].sum() else np.nan,
            '_TAUX_SUB': df['Subventions'].sum() / df['Total'].sum() if 'Subventions' in df and 'Total' in df and df['Total'].sum() else np.nan,
            '_TRLB': (df['Taux de rentabilité locative brute immédiat'] * df['Total']).sum() / df['Total'].sum() if 'Taux de rentabilité locative brute immédiat' in df and 'Total' in df and df['Total'].sum() else np.nan,
            '_EBE': (df['EBE / Loyers potentiels Cumulés'] * df['Montant']).sum() / df['Montant'].sum() if 'EBE / Loyers potentiels Cumulés' in df and 'Montant' in df and df['Montant'].sum() else np.nan,
            '_DETTE': (df['Service de la Dette / Loyers potentiels Cumulés'] * df['Montant']).sum() / df['Montant'].sum() if 'Service de la Dette / Loyers potentiels Cumulés' in df and 'Montant' in df and df['Montant'].sum() else np.nan,
            '_AUTOFI': (df["Taux d'Autofinancement"] * df['Montant']).sum() / df['Montant'].sum() if "Taux d'Autofinancement" in df and 'Montant' in df and df['Montant'].sum() else np.nan,
        }
        results.append(total)
        return pd.DataFrame(results)

    def parse_all(self):
        print("parse_all: called")
        data = defaultdict(lambda: pd.DataFrame(columns=['Key', 'Value']))
        data.update(self.parse_id())
        synth = self.parse_synth()
        
        print("parse_all: synth type:", type(synth), "shape:", getattr(synth, 'shape', None))
        print("parse_all: synth columns:", getattr(synth, 'columns', None))
        print("parse_all: synth head:\n", getattr(synth, 'head', lambda: None)())

        if not synth.empty:
            # On suppose que synth est un DataFrame avec les colonnes nécessaires
            data['SYNTHESE'] = self.compute_synthese(synth)
            print("parse_all: SYNTHESE shape:", getattr(data['SYNTHESE'], 'shape', None))
        loc = self.parse_locatif()
        if not loc.empty:
            data['ANALYSE_LOYERS'] = loc
        prp = self.parse_prp()
        if not prp.empty:
            data['ANALYSE_PRP'] = prp
        finan = self.parse_financement()
        if not finan.empty:
            data['ANALYSE_FINANCEMENT'] = finan
        print("parse_all: keys:", list(data.keys()))
        for k, v in data.items():
            print(f"parse_all: {k} type={type(v)}, shape={getattr(v, 'shape', None)}")
        return data

    def get_fille_ids(self):
        try:
            df = pd.read_excel(self.path, sheet_name='Identif', engine='pyxlsb', header=None)
            row = df.iloc[56]  # row 57 in Excel
            ids = [c for c in row[5:10] if pd.notna(c)]
            return ids
        except Exception:
            return []
    def parse_synth(self):
        print("parse_synth: called")
        # DEBUG: Affiche la structure brute de la feuille synthèse
        try:
            df_raw = self._read_range('Fiche_Synthse', 'A1:T20')
            print("parse_synth: feuille brute A1:T20")
            print(df_raw)
        except Exception as e:
            print("parse_synth: erreur lecture brute:", e)
        data = []
        ids = self.get_fille_ids()
        print("parse_synth: ids =", ids)
        for idx, fid in enumerate(ids):
            try:
                print(f"parse_synth: reading for fid={fid}, idx={idx}")
                df = self._read_range('Fiche_Synthse', f'F18:F28')
                print("parse_synth: df loaded")
                df.columns = ['Key']
                values = self._read_range('Fiche_Synthse', f'F18:F28').iloc[:, idx]
                for k, v in zip(df['Key'], values):
                    data.append({'Id2': fid, 'Key': k, 'Value': v})
            except Exception as e:
                print(f"parse_synth: exception for fid={fid}, idx={idx}:", e)
                continue
            print("parse_synth: output columns:", df.columns)
            print("parse_synth: output head:\n", df.head())
        print("parse_synth: returning DataFrame of shape", pd.DataFrame(data).shape)
        return pd.DataFrame(data)


    def parse_locatif(self):
        """Parse locatif data from the 'LoyersEtCharges' sheet."""
        ids = self.get_fille_ids()
        if not ids:
            return pd.DataFrame()

        try:
            df = pd.read_excel(
                self.path,
                sheet_name='LoyersEtCharges',
                engine='pyxlsb',
                header=14,
                usecols='C:T',
                nrows=41,
            )
        except Exception:
            return pd.DataFrame()

        df = df.dropna(how='all')
        data = []
        for _, row in df.iterrows():
            row_id = str(row.iloc[0])
            for fid in ids:
                if str(row_id) in str(fid):
                    for col in df.columns[1:]:
                        data.append({'Id2': fid, 'Key': col, 'Value': row[col]})
        return pd.DataFrame(data)

    def parse_prp(self):
        """Parse PRP values from the 'PRP IKOS' sheet."""
        ids = self.get_fille_ids()
        if not ids:
            return pd.DataFrame()

        base_cells = ['D52', 'D80', 'D99', 'D108', 'D119', 'D124', 'D136', 'D138']
        rows = []
        for idx, fid in enumerate(ids):
            for cell in base_cells:
                try:
                    key = self._read_range('PRP IKOS', cell).iloc[0, 0]
                    val_cell = self._offset_range(cell, col_offset=10 + idx)
                    val = self._read_range('PRP IKOS', val_cell).iloc[0, 0]
                    rows.append({'Id2': fid, 'Key': key, 'Value': val})
                except Exception:
                    continue
        return pd.DataFrame(rows)

    def parse_financement(self):
        """Parse financing data from the 'Financement' sheet."""
        ids = self.get_fille_ids()
        if not ids:
            return pd.DataFrame()

        sections = [
            ('E13:E33', 4, 'Subvention'),
            ('F35:F45', 3, 'Fonds propres'),
            ('F46:F71', 3, 'Prets'),
        ]

        rows = []
        for idx, fid in enumerate(ids):
            for base_range, offset, label in sections:
                try:
                    keys = self._read_range('Financement', base_range).iloc[:, 0]
                    val_range = self._offset_range(base_range, col_offset=offset + idx)
                    values = self._read_range('Financement', val_range).iloc[:, 0]
                    for k, v in zip(keys, values):
                        rows.append({'Id2': fid, 'Key': f'{label}:{k}', 'Value': v})
                except Exception:
                    continue
        return pd.DataFrame(rows)
